import json
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_package_mapping_excel():
    """Create Excel file with three sheets showing Package Group and Specs mapping"""
    
    # Read package files
    with open('Data/elec_package.txt', 'r', encoding='utf-8') as f:
        elec_data = json.load(f)
    
    with open('Data/mech_package.txt', 'r', encoding='utf-8') as f:
        mech_data = json.load(f)
    
    with open('Data/plumbing_package.txt', 'r', encoding='utf-8') as f:
        plumbing_data = json.load(f)
    
    # Create Excel writer
    excel_file = 'Package_Group_to_Spec_Mapping.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        
        # Process Electrical
        elec_rows = []
        for package, specs in elec_data.items():
            # Join specs with newlines for better readability
            specs_str = '\n'.join(specs)
            elec_rows.append({
                'Package Group': package,
                'Specs': specs_str
            })
        
        elec_df = pd.DataFrame(elec_rows)
        elec_df.to_excel(writer, sheet_name='Electrical', index=False)
        
        # Process Mechanical
        mech_rows = []
        for package, specs in mech_data.items():
            specs_str = '\n'.join(specs)
            mech_rows.append({
                'Package Group': package,
                'Specs': specs_str
            })
        
        mech_df = pd.DataFrame(mech_rows)
        mech_df.to_excel(writer, sheet_name='Mechanical', index=False)
        
        # Process Plumbing (different format - has objects with code and title)
        plumbing_rows = []
        for package, specs in plumbing_data.items():
            # Format specs as "code - title"
            formatted_specs = []
            for spec in specs:
                if isinstance(spec, dict):
                    formatted_specs.append(f"{spec['code']} - {spec['title']}")
                else:
                    formatted_specs.append(str(spec))
            specs_str = '\n'.join(formatted_specs)
            plumbing_rows.append({
                'Package Group': package,
                'Specs': specs_str
            })
        
        plumbing_df = pd.DataFrame(plumbing_rows)
        plumbing_df.to_excel(writer, sheet_name='Plumbing', index=False)
        
        # Format all sheets
        for sheet_name in ['Electrical', 'Mechanical', 'Plumbing']:
            worksheet = writer.sheets[sheet_name]
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            # For cells with newlines, count the longest line
                            cell_lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in cell_lines)
                            if max_line_length > max_length:
                                max_length = max_line_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 30
            
            # Enable text wrapping for data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    print(f"\nExcel file '{excel_file}' created successfully!")
    print(f"Sheets created: Electrical, Mechanical, Plumbing")

if __name__ == "__main__":
    create_package_mapping_excel()

