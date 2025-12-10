import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Read the category mappings from txt files
def read_categories(txt_file):
    """Read categories from txt file and return a dictionary mapping item number to category"""
    categories = {}
    with open(txt_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        # Skip header line
        for line in lines[1:]:
            parts = line.strip().split('\t')
            if len(parts) >= 3:
                item_num = parts[0].strip()
                category = parts[2].strip()
                categories[item_num] = category
    return categories

# Read CSV files
def read_csv_data(csv_file):
    """Read CSV file and return DataFrame"""
    # Skip first two rows (Project Name and empty row)
    df = pd.read_csv(csv_file, skiprows=2, encoding='utf-8')
    # Clean column names (remove leading/trailing spaces and quotes)
    df.columns = df.columns.str.strip().str.strip('"')
    return df

# Main processing
def create_excel():
    # Define file mappings
    scope_files = {
        'Electrical': {
            'txt': 'electrical.txt',
            'csv': '26 00 00 - Electrical_BidItems.csv'
        },
        'Mechanical': {
            'txt': 'mechanical.txt',
            'csv': '23 00 00 - Mechanical_BidItems.csv'
        },
        'Plumbing': {
            'txt': 'plumbing.txt',
            'csv': '22 00 00 - Plumbing_BidItems.csv'
        }
    }
    
    # Create Excel writer
    excel_file = 'Bid_Items_By_Category.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        
        for scope_name, files in scope_files.items():
            print(f"Processing {scope_name}...")
            
            # Read categories
            categories = read_categories(files['txt'])
            
            # Read CSV data
            df = read_csv_data(files['csv'])
            
            # Extract item number from the first column (remove leading/trailing spaces and quotes)
            df['Item #'] = df.iloc[:, 0].astype(str).str.strip().str.strip('"').str.strip()
            
            # Add category column based on item number
            df['Category'] = df['Item #'].map(categories).fillna('Others')
            
            # Sort by category, then by item number
            df = df.sort_values(['Category', 'Item #'], ascending=[True, True])
            
            # Reorder columns: Item #, Bid Item Description, Category, Status, Drawing Reference, Specification Reference
            column_order = ['Item #', 'Bid Item Description', 'Category', 'Status', 'Drawing Reference', 'Specification Reference']
            # Only include columns that exist
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            # Write to Excel sheet
            df.to_excel(writer, sheet_name=scope_name, index=False)
            
            # Get the worksheet for formatting
            worksheet = writer.sheets[scope_name]
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 30
            
            # Enable text wrapping for data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    print(f"\nExcel file '{excel_file}' created successfully!")
    print(f"Sheets created: {', '.join(scope_files.keys())}")

if __name__ == "__main__":
    create_excel()

