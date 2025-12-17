import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def load_json_file(filename):
    """Load JSON file and return data"""
    if not os.path.exists(filename):
        return None
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            if not content:
                return None
            # Remove markdown code block markers if present
            if content.startswith('```json'):
                content = content[7:]  # Remove ```json
            if content.startswith('```'):
                content = content[3:]  # Remove ```
            if content.endswith('```'):
                content = content[:-3]  # Remove closing ```
            content = content.strip()
            return json.loads(content)
    except (json.JSONDecodeError, Exception) as e:
        print(f"Warning: Error loading {filename}: {e}")
        return None

def build_contract_items_map(contract_items, discipline):
    """Build a map of contract item ID to description based on discipline structure"""
    if not contract_items:
        return {}
    
    id_map = {}
    
    if discipline == 'electrical':
        # Electrical: {"1": "description", "2": "description", ...}
        id_map = {str(k): v for k, v in contract_items.items()}
    elif discipline == 'mechanical':
        # Mechanical: {"description": {"id": 1, ...}}
        for desc, item_data in contract_items.items():
            if isinstance(item_data, dict) and 'id' in item_data:
                item_id = str(item_data['id'])
                id_map[item_id] = desc  # Use the key (description) as the value
    elif discipline == 'plumbing':
        # Plumbing: {"short desc": "full desc"} - need to index by position
        # Since scope items reference numeric IDs, we'll create a 1-indexed map
        items_list = list(contract_items.items())
        for idx, (short_desc, full_desc) in enumerate(items_list, start=1):
            id_map[str(idx)] = full_desc  # Use full description
    
    return id_map

def create_grps_excel():
    """Create Excel file with scope items mapping for all MEP disciplines"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    disciplines = ['electrical', 'mechanical', 'plumbing']
    discipline_names = {
        'electrical': 'Electrical',
        'mechanical': 'Mechanical',
        'plumbing': 'Plumbing'
    }
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(vertical='top', wrap_text=True)
    
    for discipline in disciplines:
        # Load data
        scope_items_file = f'Data/grps_{discipline}_scope_items.json'
        contract_items_file = f'Data/grps_{discipline}_contract_items.json'
        
        scope_items = load_json_file(scope_items_file)
        contract_items_raw = load_json_file(contract_items_file)
        
        # Build contract items map based on discipline structure
        contract_items = build_contract_items_map(contract_items_raw, discipline)
        
        if not scope_items:
            print(f"Warning: {scope_items_file} not found, skipping {discipline}")
            continue
        
        # Create worksheet
        ws = wb.create_sheet(title=discipline_names[discipline])
        
        # Set column headers
        headers = ['Scope Item', 'Contract Items (Derived From)']
        ws.append(headers)
        
        # Style header row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        
        # Set column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 80
        
        # Add data rows
        row_num = 2
        for scope_item_name, scope_data in scope_items.items():
            scope_item_id = scope_data.get('scope_item_id', '')
            combined_from = scope_data.get('combined_from', [])
            
            # Format scope item name with ID
            scope_item_display = f"{scope_item_name}"
            if scope_item_id:
                scope_item_display = f"[{scope_item_id}] {scope_item_name}"
            
            # Get contract item descriptions
            contract_item_list = []
            for contract_id in combined_from:
                contract_id_str = str(contract_id)
                if contract_items and contract_id_str in contract_items:
                    contract_desc = contract_items[contract_id_str]
                    contract_item_list.append(f"{contract_id}: {contract_desc}")
                else:
                    contract_item_list.append(f"{contract_id}: (Not found in contract items)")
            
            contract_items_text = '\n'.join(contract_item_list) if contract_item_list else 'None'
            
            # Add row
            ws.cell(row=row_num, column=1, value=scope_item_display)
            ws.cell(row=row_num, column=2, value=contract_items_text)
            
            # Style cells
            for col in range(1, 3):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border_style
                cell.alignment = wrap_align if col == 2 else Alignment(vertical='top')
            
            row_num += 1
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # Save workbook to Data folder
    output_file = 'Data/GRPS_Scope_Items_Mapping.xlsx'
    try:
        wb.save(output_file)
        print(f"Excel file '{output_file}' created successfully!")
        print(f"Sheets created: {', '.join(wb.sheetnames)}")
    except PermissionError:
        # Try with a timestamp if file is open
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f'Data/GRPS_Scope_Items_Mapping_{timestamp}.xlsx'
        wb.save(output_file)
        print(f"Excel file '{output_file}' created successfully! (Original file was open)")
        print(f"Sheets created: {', '.join(wb.sheetnames)}")
    
    return output_file

if __name__ == "__main__":
    create_grps_excel()

